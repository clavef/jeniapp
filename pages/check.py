# pages/check.py
import streamlit as st
st.set_page_config(page_title="제니앱", page_icon="📱", layout="wide")
import json
import pandas as pd
import zipfile
import io
import datetime
from io import BytesIO
from shared import show_menu
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

show_menu("인스타 언팔체크")

st.title("📱 인스타 언팔체크")

st.markdown("""
**언팔로워란?**

내가 팔로우하고 있지만 상대는 나를 팔로우하지 않는 계정을 뜻합니다.  
즉, 일방적으로 팔로우 중인 사람(맞팔이 아닌 관계)이죠.

**⛳️ 사용 안내**
- 이 도구는 **PC와 모바일 모두 사용 가능**하지만 **크롬(Chrome)** 브라우저에 최적화되어 있습니다.
- 네이버앱이나 Safari 등의 브라우저에서는 일부 기능이 제대로 작동하지 않을 수 있습니다.

---

### 📁 인스타그램 ZIP 파일 다운로드 방법

1️⃣ **모바일**: 내 인스타그램 프로필 → 오른쪽 상단 ☰ 아이콘 클릭  
   **PC**: 인스타그램 웹 로그인 → 왼쪽 하단 "더 보기" 클릭

2️⃣ **내 활동** → **내 정보 다운로드** → 계속

3️⃣ **다운로드 요청** → 계정 선택 후 다음

4️⃣ '전체 사본' 대신 '정보 유형 선택'을 선택 → **팔로워 및 팔로잉만 체크** 후 다음

5️⃣ **전체 기간**, 형식은 **JSON**으로 설정 후 요청 제출

6️⃣ 이메일 수신 후 다운로드 (소요 시간은 데이터 양에 따라 다름)

7️⃣ 받은 **ZIP 파일을 아래 업로드**하면 분석이 자동으로 시작됩니다

---

### 🔎 사용 방법
1. 위 절차대로 받은 **ZIP 파일을 업로드**하세요.  
2. 업로드된 파일 내에서 팔로워 및 팔로잉 정보를 자동으로 추출합니다.  
3. 내가 팔로우하지만 나를 팔로우하지 않는 계정을 **정렬된 표로 확인**하고,  
   **엑셀 파일로 다운로드**할 수 있습니다. (하이퍼링크 포함)

"""
)

uploaded_zip = st.file_uploader("인스타그램 ZIP 파일 업로드", type="zip")


def extract_following_info(data):
    results = []
    if isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list):
                for entry in v:
                    if "string_list_data" in entry:
                        string_data = entry["string_list_data"][0]
                        username = string_data.get("value")
                        timestamp = string_data.get("timestamp")
                        results.append({"username": username, "timestamp": timestamp})
    elif isinstance(data, list):
        for entry in data:
            if "string_list_data" in entry:
                string_data = entry["string_list_data"][0]
                username = string_data.get("value")
                timestamp = string_data.get("timestamp")
                results.append({"username": username, "timestamp": timestamp})
    return results

def format_time(ts):
    if not ts:
        return "-"
    dt = datetime.datetime.fromtimestamp(ts)
    delta_days = (datetime.datetime.now() - dt).days
    formatted = dt.strftime("%Y.%m.%d %H:%M")
    return f"{delta_days}일 전, {formatted}"

def find_json_file(zip_file, keyword):
    files = [f for f in zip_file.namelist() if keyword in f and f.endswith(".json")]
    if keyword == "followers":
        files = [f for f in files if "followers_1.json" in f]
    elif keyword == "following":
        files = [f for f in files if f.endswith("following.json")]
    return files[0] if files else None

if uploaded_zip:
    try:
        with zipfile.ZipFile(uploaded_zip) as z:
            followers_file = find_json_file(z, "followers")
            following_file = find_json_file(z, "following")

            if not followers_file or not following_file:
                st.error("ZIP 파일에서 followers 또는 following JSON 파일을 찾을 수 없습니다.")
            else:
                with z.open(followers_file) as f:
                    followers_data = json.load(f)
                with z.open(following_file) as f:
                    following_data = json.load(f)

                follower_usernames = set([entry["username"] for entry in extract_following_info(followers_data)])
                following_info = extract_following_info(following_data)

                results = []
                for entry in following_info:
                    username = entry["username"]
                    timestamp = entry["timestamp"]
                    if username not in follower_usernames:
                        results.append({
                            "ID": f"@{username}",
                            "링크": f"https://instagram.com/{username}",
                            "내가 팔로잉한 날짜": format_time(timestamp),
                            "timestamp_raw": timestamp or 0
                        })

                st.success(f"총 {len(results)}명이 나를 팔로우하지 않아요.")

                # 정렬 옵션
                sort_order = st.radio("정렬 순서 선택", ["최신순", "오래된순"], horizontal=True)
                results = sorted(results, key=lambda x: x["timestamp_raw"], reverse=(sort_order == "최신순"))

                # 웹용 테이블 출력
                display_df = pd.DataFrame(results)
                display_df["ID"] = display_df.apply(
                    lambda row: f'<a href="{row["링크"]}" target="_blank">{row["ID"]}</a>', axis=1
                )
                st.write("#### 결과:", unsafe_allow_html=True)
                st.write(
                    display_df[["ID", "내가 팔로잉한 날짜"]]
                    .to_html(escape=False, index=False, justify="left"),
                    unsafe_allow_html=True
                )

                # XLSX 다운로드 (하이퍼링크 포함)
                df_export = pd.DataFrame(results)[["ID", "링크", "내가 팔로잉한 날짜"]]
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Unfollow Check"

                ws.append(["ID", "내가 팔로잉한 날짜"])

                for i in range(len(df_export)):
                    id_cell = ws.cell(row=i+2, column=1, value=df_export.iloc[i]["ID"])
                    id_cell.hyperlink = df_export.iloc[i]["링크"]
                    id_cell.style = "Hyperlink"
                    date_cell = ws.cell(row=i+2, column=2, value=df_export.iloc[i]["내가 팔로잉한 날짜"])

                for col in ws.columns:
                    for cell in col:
                        cell.alignment = Alignment(horizontal="left")

                output = BytesIO()
                wb.save(output)
                st.download_button(
                    label="📅 엑셀파일 다운로드",
                    data=output.getvalue(),
                    file_name="제니앱_인스타_언팔체크.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.error(f"처리 중 오류 발생: {e}")
